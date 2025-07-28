import streamlit as st
import pandas as pd
import pdfplumber
import re
import openpyxl
import streamlit as st
import pandas as pd
import pdfplumber
import re
import tempfile
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime
import tempfile
import os
import shutil

# Konfigurasi halaman
st.set_page_config(
    page_title="Aplikasi Penerimaan Rusun",
    page_icon="🏢",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Inisialisasi session state
def init_session_state():
    """Inisialisasi session state untuk menyimpan data antar tahapan"""
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'df_bank' not in st.session_state:
        st.session_state.df_bank = None
    if 'df_setortunai' not in st.session_state:
        st.session_state.df_setortunai = None
    if 'df_non_rusun' not in st.session_state:
        st.session_state.df_non_rusun = None
    if 'df_final' not in st.session_state:
        st.session_state.df_final = None
    if 'results' not in st.session_state:
        st.session_state.results = None
    if 'valid_data' not in st.session_state:
        st.session_state.valid_data = None
    if 'export_file' not in st.session_state:
        st.session_state.export_file = None

# Fungsi utilitas
def save_uploadedfile_temp(uploaded_file, subfolder=""):
    """Menyimpan file yang diupload ke folder temporary"""
    temp_dir = os.path.join(tempfile.gettempdir(), "rusun_app", subfolder)
    os.makedirs(temp_dir, exist_ok=True)
    
    file_path = os.path.join(temp_dir, uploaded_file.name)
    with open(file_path, "wb") as f:
        f.write(uploaded_file.getbuffer())
    
    return file_path

def to_numeric_safe(value):
    """Konversi nilai ke numeric dengan aman"""
    try:
        if pd.isna(value) or value == '' or value == 'None':
            return 0
        numeric_value = float(str(value).replace(',', ''))
        return numeric_value
    except:
        return 0

def convert_to_first_of_month_date(bulan, tahun):
    """Mengonversi string bulan dan tahun menjadi objek datetime tanggal 1."""
    try:
        if pd.isna(bulan) or pd.isna(tahun) or bulan == '' or tahun == '':
            return None # Mengembalikan nilai kosong yang dikenali pandas
        
        month_map = {
            'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4, 'Mei': 5, 'Juni': 6,
            'Juli': 7, 'Agustus': 8, 'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
        }
        month_num = month_map.get(str(bulan))
        
        if month_num and str(tahun).isdigit():
            # Membuat objek tanggal asli: tanggal 1, di bulan & tahun yang sesuai
            return datetime(int(tahun), month_num, 1)
        return None
    except:
        return None

# Fungsi pemrosesan PDF (dari notebook)
def process_pdf(pdf_path):
    """Mengekstrak data dari PDF Bank Statement (sama persis dengan test3.ipynb cell 1)"""
    all_text = ''
    all_tables = []
    header = None

    # Buka file PDF
    with pdfplumber.open(pdf_path) as pdf:
        # Loop untuk setiap halaman di dalam PDF
        for page_num, page in enumerate(pdf.pages):
            # Ekstrak semua tabel dari halaman tersebut
            tables_on_page = page.extract_tables()
            
            for table in tables_on_page:
                if table:  # Pastikan tabel tidak kosong
                    if page_num == 0:  # Halaman pertama
                        # Ambil header dari halaman pertama
                        if header is None:
                            header = table[0]
                        # Ambil semua data setelah header
                        all_tables.extend(table[1:])
                    else:  # Halaman kedua dan seterusnya
                        # Periksa apakah baris pertama adalah header yang sama
                        if len(table) > 0:
                            first_row = table[0]
                            # Bandingkan dengan header yang sudah ada
                            if header and first_row == header:
                                # Skip header dan ambil data mulai dari baris kedua
                                if len(table) > 1:
                                    all_tables.extend(table[1:])
                            else:
                                # Jika baris pertama bukan header yang sama, anggap sebagai data
                                all_tables.extend(table)

    # Buat DataFrame dengan header yang sudah diambil dari halaman pertama
    if header and all_tables:
        df_bank = pd.DataFrame(all_tables, columns=header)
        
        # Cari kolom yang kemungkinan berisi narasi
        narasi_keywords = ['Narasi']
        narasi_columns = [col for col in df_bank.columns if any(keyword in str(col).lower() for keyword in narasi_keywords)]
        
        if narasi_columns:
            for col in narasi_columns:
                # Hapus semua \n dan \r dari kolom narasi
                df_bank[col] = df_bank[col].astype(str).str.replace('\n', ' ', regex=False).str.replace('\r', ' ', regex=False)
                # Bersihkan juga spasi berlebih
                df_bank[col] = df_bank[col].str.replace(r'\s+', '', regex=True).str.strip()
        else:
            # Membersihkan semua kolom string dari karakter newline
            for col in df_bank.columns:
                if df_bank[col].dtype == 'object':  # Kolom string
                    df_bank[col] = df_bank[col].astype(str).str.replace('\n', ' ', regex=False).str.replace('\r', ' ', regex=False)
                    df_bank[col] = df_bank[col].str.replace(r'\s+', '', regex=True).str.strip()
        
        # PEMBERSIHAN KOLOM CREDIT TRANSACTION DAN BALANCE
        
        # Fungsi untuk membersihkan format angka (hilangkan 2 angka di belakang koma dan koma)
        def clean_numeric_column(value):
            if pd.isna(value) or value == '' or value == 'nan':
                return value
            
            value_str = str(value).strip()
            
            # Jika ada titik desimal, hilangkan 2 angka terakhir (desimal)
            if '.' in value_str:
                # Pisahkan bagian sebelum dan sesudah titik
                parts = value_str.split('.')
                if len(parts) == 2 and len(parts[1]) >= 2:
                    # Hilangkan 2 angka terakhir (desimal)
                    value_str = parts[0]
            
            # Hilangkan semua koma
            value_str = value_str.replace(',', '')
            
            return value_str
        
        # Daftar kolom yang perlu dibersihkan
        columns_to_clean = ['Credit Transaction', 'Balance']
        
        # Cek dan bersihkan kolom yang ada
        for col_name in columns_to_clean:
            # Cari kolom yang mengandung nama tersebut (case insensitive)
            matching_cols = [col for col in df_bank.columns if col_name.lower() in str(col).lower()]
            
            if matching_cols:
                for col in matching_cols:
                    # Terapkan pembersihan
                    df_bank[col] = df_bank[col].apply(clean_numeric_column)
        
        # PEMBERSIHAN DAN FORMAT KOLOM TANGGAL
        
        # Fungsi untuk memformat tanggal menjadi mm/dd/yyyy
        def format_date_column(value):
            """
            Memformat kolom tanggal menjadi format mm/dd/yyyy
            Menangani berbagai format input tanggal termasuk format DDMmmYYYY
            """
            if pd.isna(value) or value == '' or value == 'nan':
                return value
            
            value_str = str(value).strip()
            
            # Jika sudah dalam format mm/dd/yyyy, kembalikan
            if re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', value_str):
                # Parse dan validasi format yang ada
                try:
                    # Coba parse sebagai mm/dd/yyyy
                    parsed_date = datetime.strptime(value_str, '%m/%d/%Y')
                    return parsed_date.strftime('%m/%d/%Y')
                except ValueError:
                    # Jika gagal, mungkin format dd/mm/yyyy, lanjut ke parsing lain
                    pass
            
            # Handle format DDMmmYYYY (seperti 03Jun2025)
            month_mapping = {
                'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06',
                'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12',
                'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 'MAY': '05', 'JUN': '06',
                'JUL': '07', 'AUG': '08', 'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'
            }
            
            # Pattern untuk DDMmmYYYY (03Jun2025)
            ddmmmyyyy_pattern = r'(\d{2})([A-Za-z]{3})(\d{4})'
            match = re.match(ddmmmyyyy_pattern, value_str)
            if match:
                day, month_str, year = match.groups()
                month_num = month_mapping.get(month_str)
                if month_num:
                    return f"{month_num}/{day}/{year}"  # Format mm/dd/yyyy
            
            # Coba parse berbagai format tanggal
            date_formats = [
                '%d/%m/%Y',    # 13/06/2025
                '%d-%m-%Y',    # 13-06-2025  
                '%Y-%m-%d',    # 2025-06-13
                '%d/%m/%y',    # 13/06/25
                '%d-%m-%y',    # 13-06-25
                '%d %m %Y',    # 13 06 2025
                '%d.%m.%Y',    # 13.06.2025
                '%d%b%Y',      # 13Jun2025
                '%d-%b-%Y',    # 13-Jun-2025
                '%m/%d/%Y',    # 06/13/2025 (format US)
                '%m-%d-%Y',    # 06-13-2025 (format US)
            ]
            
            for fmt in date_formats:
                try:
                    # Parse tanggal dengan format yang dicoba
                    parsed_date = datetime.strptime(value_str, fmt)
                    # Return dalam format mm/dd/yyyy
                    return parsed_date.strftime('%m/%d/%Y')
                except ValueError:
                    continue
            
            # Jika tidak bisa di-parse, coba ekstrak angka dan format ulang
            # Pattern untuk mencari tanggal dalam berbagai format
            date_patterns = [
                r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',  # dd/mm/yyyy atau dd-mm-yyyy
                r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',  # yyyy/mm/dd atau yyyy-mm-dd
                r'(\d{1,2})[/-](\d{1,2})[/-](\d{2})',  # dd/mm/yy atau dd-mm-yy
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, value_str)
                if match:
                    groups = match.groups()
                    
                    if len(groups[2]) == 4:  # Format dengan tahun 4 digit
                        if len(groups[0]) == 4:  # yyyy/mm/dd
                            year, month, day = groups[0], groups[1], groups[2]
                        else:  # dd/mm/yyyy (asumsi format Indonesia)
                            day, month, year = groups[0], groups[1], groups[2]
                    else:  # Format dengan tahun 2 digit
                        day, month, year = groups[0], groups[1], groups[2]
                        # Konversi tahun 2 digit ke 4 digit
                        year_int = int(year)
                        if year_int >= 20 and year_int <= 30:
                            year = '20' + year
                        else:
                            year = '20' + year  # Default ke 20xx
                    
                    # Pastikan format dengan leading zero
                    day = day.zfill(2)
                    month = month.zfill(2)
                    
                    # Validasi tanggal
                    try:
                        datetime.strptime(f"{day}/{month}/{year}", '%d/%m/%Y')
                        return f"{month}/{day}/{year}"  # Return format mm/dd/yyyy
                    except ValueError:
                        continue
            
            # Jika semua gagal, kembalikan nilai asli
            return value_str
        
        # Daftar kolom tanggal yang perlu diformat
        date_columns = ['Posting Date', 'Effective Date']
        
        # Format kolom tanggal
        for col_name in date_columns:
            # Cari kolom yang mengandung nama tersebut (case insensitive)
            matching_cols = [col for col in df_bank.columns if col_name.lower() in str(col).lower()]
            
            if matching_cols:
                for col in matching_cols:
                    print(f"📅 Memformat kolom tanggal: {col} ke format mm/dd/yyyy")
                    # Terapkan format tanggal
                    df_bank[col] = df_bank[col].apply(format_date_column)
        
        return df_bank
    else:
        return pd.DataFrame()

def process_setortunai(df_bank):
    """Memproses data SETORTUNAI dari dataframe bank (sama persis dengan test3.ipynb cell 2)"""
    # Pisahkan data berdasarkan kata SETORTUNAI
    df_bank['Has_SETORTUNAI'] = df_bank['Narasi'].str.upper().str.contains('SETORTUNAI', na=False)

    # Data dengan SETORTUNAI (akan diproses untuk ekstraksi)
    df_setortunai = df_bank[df_bank['Has_SETORTUNAI'] == True].copy()

    # Data tanpa SETORTUNAI (akan disimpan sebagai non-rusun)
    df_non_rusun = df_bank[df_bank['Has_SETORTUNAI'] == False].copy()

    # Fungsi ekstraksi yang diperbaiki berdasarkan analisis contoh narasi
    def extract_setortunai_only(narasi_text):
        """
        Fungsi ekstraksi yang diperbaiki untuk menangani pattern NOV25, DES24, dll
        dan tidak memberikan default bulan/tahun jika tidak ada indikasi yang jelas
        """
        if pd.isna(narasi_text) or narasi_text == '' or narasi_text == 'nan':
            return '', '', ''
        
        narasi_str = str(narasi_text).upper().strip()
        
        # Hanya proses jika mengandung SETORTUNAI
        if 'SETORTUNAI' not in narasi_str:
            return '', '', ''
        
        # === KODE 8 DIGIT ===
        kode_8_digit = ''
        
        # Cari semua 8 digit dalam narasi dengan pattern yang lebih spesifik
        # 2 angka awal harus 01, 02, atau 03
        specific_8_digits = re.findall(r'\b(0[123]\d{6})\b', narasi_str)
        
        if specific_8_digits:
            # Ambil yang pertama ditemukan yang sesuai kriteria
            kode_8_digit = specific_8_digits[0]
        else:
            # Fallback: cari semua 8 digit dan filter yang dimulai 01, 02, 03
            all_8_digits = re.findall(r'\d{8}', narasi_str)
            
            for code in all_8_digits:
                # Filter kode yang valid: dimulai dengan 01, 02, atau 03 dan bukan tahun
                if (code.startswith(('01', '02', '03')) and 
                    not code.startswith(('2024', '2025', '2026'))):
                    kode_8_digit = code
                    break
        
        # === ANALISIS 40 HURUF TERAKHIR ===
        last_40_chars = narasi_str[-40:] if len(narasi_str) >= 40 else narasi_str
        
        # === BULAN - PERBAIKAN DENGAN PRIORITAS YANG TEPAT ===
        bulan = ''
        
        # 1. PRIORITAS TERTINGGI: Cari nama bulan lengkap + tahun 4 digit di 40 huruf terakhir
        full_month_year_4digit = r'(JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)(20\d{2})'
        matches = re.findall(full_month_year_4digit, last_40_chars)
        if matches:
            month_name = matches[-1][0]  # Ambil nama bulan
            month_dict = {
                'JANUARI': 'Januari', 'FEBRUARI': 'Februari', 'MARET': 'Maret', 'APRIL': 'April',
                'MEI': 'Mei', 'JUNI': 'Juni', 'JULI': 'Juli', 'AGUSTUS': 'Agustus',
                'SEPTEMBER': 'September', 'OKTOBER': 'Oktober', 'NOVEMBER': 'November', 'DESEMBER': 'Desember'
            }
            bulan = month_dict.get(month_name, '')
        
        # 2. PRIORITAS KEDUA: Cari pattern bulan singkat 3 huruf + tahun 4 digit (FEB2025, JAN2025, DES2024)
        if not bulan:
            short_month_year_4digit = r'(JAN|FEB|MAR|APR|MEI|JUN|JUL|AGS|AGST|SEP|SEPT|OKT|NOV|DES)(20\d{2})'
            matches = re.findall(short_month_year_4digit, last_40_chars)
            if matches:
                month_abbr = matches[-1][0]  # Ambil yang terakhir ditemukan
                short_month_dict = {
                    'JAN': 'Januari', 'FEB': 'Februari', 'MAR': 'Maret', 'APR': 'April',
                    'MEI': 'Mei', 'JUN': 'Juni', 'JUL': 'Juli', 'AGS': 'Agustus', 'AGST': 'Agustus',
                    'SEP': 'September', 'SEPT': 'September', 'OKT': 'Oktober', 'NOV': 'November', 'DES': 'Desember'
                }
                bulan = short_month_dict.get(month_abbr, '')
        
        # 3. Jika belum ada, cari nama bulan lengkap di seluruh narasi
        if not bulan:
            full_month_names = ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI', 
                               'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER']
            
            for month_name in full_month_names:
                if month_name in narasi_str:
                    month_dict = {
                        'JANUARI': 'Januari', 'FEBRUARI': 'Februari', 'MARET': 'Maret', 'APRIL': 'April',
                        'MEI': 'Mei', 'JUNI': 'Juni', 'JULI': 'Juli', 'AGUSTUS': 'Agustus',
                        'SEPTEMBER': 'September', 'OKTOBER': 'Oktober', 'NOVEMBER': 'November', 'DESEMBER': 'Desember'
                    }
                    bulan = month_dict[month_name]
                    break
        
        # 4. Cari pattern bulan singkat + 2 digit tahun (NOV25, DES24, dll)
        if not bulan:
            month_year_short = {
                'JAN': 'Januari', 'FEB': 'Februari', 'MAR': 'Maret', 'APR': 'April',
                'MEI': 'Mei', 'JUN': 'Juni', 'JUL': 'Juli', 'AGS': 'Agustus', 'AGST': 'Agustus',
                'SEP': 'September', 'SEPT': 'September', 'OKT': 'Oktober', 'NOV': 'November', 'DES': 'Desember'
            }
            
            # Pattern: 3 huruf bulan + 2 digit tahun
            month_pattern = r'(JAN|FEB|MAR|APR|MEI|JUN|JUL|AGS|AGST|SEP|SEPT|OKT|NOV|DES)(\d{2})'
            matches = re.findall(month_pattern, last_40_chars)
            
            if matches:
                month_abbr = matches[-1][0]  # Ambil yang terakhir ditemukan
                bulan = month_year_short.get(month_abbr, '')
        
        # 5. Cari pattern BLN + singkatan (existing)
        if not bulan:
            bln_abbreviations = {
                'BLNJAN': 'Januari', 'BLNFEB': 'Februari', 'BLNMAR': 'Maret', 'BLNAPR': 'April',
                'BLNMEI': 'Mei', 'BLNJUN': 'Juni', 'BLNJUL': 'Juli', 'BLNAGST': 'Agustus',
                'BLNSEPT': 'September', 'BLNOKT': 'Oktober', 'BLNNOV': 'November', 'BLNDES': 'Desember'
            }
            
            for bln_code, month_name in bln_abbreviations.items():
                if bln_code in last_40_chars:
                    bulan = month_name
                    break
        
        # 6. Jika masih belum ada, cari angka bulan di 40 huruf terakhir (tetapi lebih selektif)
        if not bulan:
            # Pattern untuk angka bulan dalam konteks yang jelas
            month_patterns = [
                r'/(\d{1,2})/',        # /MM/ atau /M/
                r'/(\d{1,2})\b',       # /MM atau /M (di akhir)
            ]
            
            for pattern in month_patterns:
                matches = re.findall(pattern, last_40_chars)
                for match in matches:
                    if match.isdigit() and 1 <= int(match) <= 12:
                        month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                                     'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
                        bulan = month_names[int(match)]
                        break
                if bulan:
                    break
        
        # === TAHUN - PERBAIKAN DENGAN PRIORITAS YANG TEPAT ===
        tahun = ''
        
        # 1. PRIORITAS TERTINGGI: Cari format bulan lengkap + tahun 4 digit DULU (MEI2024, DES2024)
        full_month_year_4digit = r'(JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)(20\d{2})'
        matches = re.findall(full_month_year_4digit, last_40_chars)
        if matches:
            year_part = matches[-1][-1]  # Ambil bagian tahun (20XX)
            if year_part.isdigit() and len(year_part) == 4:
                tahun = year_part
        
        # 2. PRIORITAS KEDUA: Cari pattern bulan singkat 3 huruf + tahun 4 digit (FEB2025, JAN2025, DES2024)
        if not tahun:
            short_month_year_4digit = r'(JAN|FEB|MAR|APR|MEI|JUN|JUL|AGS|AGST|SEP|SEPT|OKT|NOV|DES)(20\d{2})'
            matches = re.findall(short_month_year_4digit, last_40_chars)
            if matches:
                year_part = matches[-1][-1]  # Ambil bagian tahun (20XX)
                if year_part.isdigit() and len(year_part) == 4:
                    tahun = year_part
        
        # 3. Jika belum ada, cari pattern bulan singkat + 2 digit tahun (NOV25, DES24)
        if not tahun:
            month_pattern = r'(JAN|FEB|MAR|APR|MEI|JUN|JUL|AGS|AGST|SEP|SEPT|OKT|NOV|DES)(\d{2})'
            matches = re.findall(month_pattern, last_40_chars)
            
            if matches:
                year_part = matches[-1][1]  # Ambil bagian tahun
                if year_part.isdigit() and len(year_part) == 2:
                    # Konversi 24->2024, 25->2025, dll
                    if int(year_part) >= 20 and int(year_part) <= 30:  # Range masuk akal
                        tahun = '20' + year_part
        
        # 4. Cari format bulan lengkap + tahun 2 digit (jika ada)
        if not tahun:
            full_month_year_2digit = r'(JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)(\d{2})'
            matches = re.findall(full_month_year_2digit, last_40_chars)
            if matches:
                year_part = matches[-1][-1]  # Ambil bagian tahun
                if year_part.isdigit() and len(year_part) == 2:
                    tahun = '20' + year_part
        
        # 5. Cari pattern BLN + bulan + tahun
        if not tahun:
            bln_pattern = r'BLN(JAN|FEB|MAR|APR|MEI|JUN|JUL|AGST|SEPT|OKT|NOV|DES)(\d{2})'
            matches = re.findall(bln_pattern, last_40_chars)
            if matches:
                year_part = matches[-1][1]
                if year_part.isdigit() and len(year_part) == 2:
                    tahun = '20' + year_part
        
        # 6. Cari tahun 4 digit standalone
        if not tahun:
            year_4_digit = re.findall(r'\b(20[0-9]{2})\b', last_40_chars)
            if year_4_digit:
                tahun = year_4_digit[-1]
        
        # PENTING: Tidak memberikan default tahun jika tidak ada indikasi yang jelas
        # Berdasarkan analisis, tidak semua narasi memiliki tahun yang valid
        
        return kode_8_digit, bulan, tahun

    # Proses data SETORTUNAI
    if len(df_setortunai) > 0:
        # Reset kolom ekstraksi untuk data SETORTUNAI
        extraction_results = df_setortunai['Narasi'].apply(extract_setortunai_only)
        
        df_setortunai['Kode_8_Digit'] = extraction_results.apply(lambda x: x[0])
        df_setortunai['Bulan'] = extraction_results.apply(lambda x: x[1])
        df_setortunai['Tahun'] = extraction_results.apply(lambda x: x[2])
        
        # Hapus kolom helper
        df_setortunai = df_setortunai.drop('Has_SETORTUNAI', axis=1)

    # Proses data NON-RUSUN
    if len(df_non_rusun) > 0:
        # Untuk data non-rusun, kosongkan kolom ekstraksi
        df_non_rusun['Kode_8_Digit'] = ''
        df_non_rusun['Bulan'] = ''
        df_non_rusun['Tahun'] = ''
        
        # Hapus kolom helper
        df_non_rusun = df_non_rusun.drop('Has_SETORTUNAI', axis=1)
    
    return df_setortunai, df_non_rusun

def filter_incomplete_data(df_setortunai, df_non_rusun):
    """Filter data yang tidak lengkap dan tahun yang tidak didukung"""
    # Cek data yang lengkap
    incomplete_mask = (
        (df_setortunai['Kode_8_Digit'] == '') |
        (df_setortunai['Bulan'] == '') |
        (df_setortunai['Tahun'] == '')
    )
    
    # Cek data dengan tahun yang tidak didukung (selain 2024 dan 2025)
    unsupported_year_mask = ~df_setortunai['Tahun'].astype(str).isin(['2024', '2025'])
    
    # Cek data dengan Credit Transaction > 600.000
    high_credit_mask = df_setortunai['Credit Transaction'].apply(to_numeric_safe) > 600000
    
    # Gabungkan ketiga filter: data tidak lengkap ATAU tahun tidak didukung ATAU credit transaction > 600.000
    to_move_mask = incomplete_mask | unsupported_year_mask | high_credit_mask
    
    # Pindahkan data tidak lengkap dan tahun tidak didukung ke non-rusun
    df_setortunai_to_move = df_setortunai[to_move_mask].copy()
    df_setortunai_complete = df_setortunai[~to_move_mask].copy()
    
    # Tambahkan keterangan untuk data yang dipindah
    def create_keterangan(row):
        missing_fields = []
        reasons = []
        
        # Cek kelengkapan data
        if row['Kode_8_Digit'] == '':
            missing_fields.append('Kode 8 Digit')
        if row['Bulan'] == '':
            missing_fields.append('Bulan')
        if row['Tahun'] == '':
            missing_fields.append('Tahun')
        
        # Cek tahun yang tidak didukung
        if str(row['Tahun']) not in ['2024', '2025'] and str(row['Tahun']) != '':
            reasons.append(f'Tahun {row["Tahun"]} tidak didukung (hanya 2024-2025)')
        
        # Cek Credit Transaction > 600.000
        credit_amount = to_numeric_safe(row.get('Credit Transaction', 0))
        if credit_amount > 600000:
            reasons.append(f'Credit Transaction {credit_amount:,.0f} > 600.000 (perlu cek manual)')
        
        # Gabungkan alasan
        if missing_fields:
            reasons.append(f"Kekurangan: {', '.join(missing_fields)}")
        
        return ' | '.join(reasons) if reasons else 'Data tidak lengkap'
    
    df_setortunai_to_move['Keterangan'] = df_setortunai_to_move.apply(create_keterangan, axis=1)
    
    # Gabungkan dengan data non-rusun
    df_non_rusun_new = pd.concat([df_non_rusun, df_setortunai_to_move], ignore_index=True)
    
    return df_setortunai_complete, df_non_rusun_new

def extract_from_master_excel(df_setortunai, master_files):
    """Ekstrak data dari Master Excel (disesuaikan dengan kode dari notebook) - OPTIMIZED VERSION"""
    import openpyxl
    import pandas as pd
    import streamlit as st
    from datetime import datetime
    
    # Mapping kolom extract per bulan (4 kolom: Nama Penghuni, Tanggal Perjanjian, Sewa Hunian, Sewa Lahan)
    month_mappings_extract = [
        ('Januari', ['I', 'J', 'K', 'L']),
        ('Februari', ['S', 'T', 'U', 'V']),
        ('Maret', ['AC', 'AD', 'AE', 'AF']),
        ('April', ['AM', 'AN', 'AO', 'AP']),
        ('Mei', ['AW', 'AX', 'AY', 'AZ']),
        ('Juni', ['BG', 'BH', 'BI', 'BJ']),
        ('Juli', ['BQ', 'BR', 'BS', 'BT']),
        ('Agustus', ['CA', 'CB', 'CC', 'CD']),
        ('September', ['CK', 'CL', 'CM', 'CN']),
        ('Oktober', ['CU', 'CV', 'CW', 'CX']),
        ('November', ['DE', 'DF', 'DG', 'DH']),
        ('Desember', ['DO', 'DP', 'DQ', 'DR'])
    ]
    
    extract_cols = ['Nama Penghuni', 'Tanggal Perjanjian Sewa', 'Sewa Hunian', 'Sewa Lahan Lantai 1']
    
    def get_month_columns_extract(bulan):
        """Mengembalikan kolom berdasarkan bulan"""
        month_mapping = {
            'Januari': ['I', 'J', 'K', 'L'],
            'Februari': ['S', 'T', 'U', 'V'],
            'Maret': ['AC', 'AD', 'AE', 'AF'],
            'April': ['AM', 'AN', 'AO', 'AP'],
            'Mei': ['AW', 'AX', 'AY', 'AZ'],
            'Juni': ['BG', 'BH', 'BI', 'BJ'],
            'Juli': ['BQ', 'BR', 'BS', 'BT'],
            'Agustus': ['CA', 'CB', 'CC', 'CD'],
            'September': ['CK', 'CL', 'CM', 'CN'],
            'Oktober': ['CU', 'CV', 'CW', 'CX'],
            'November': ['DE', 'DF', 'DG', 'DH'],
            'Desember': ['DO', 'DP', 'DQ', 'DR']
        }
        return month_mapping.get(bulan, None)
    
    def get_sheet_name_extract(kode_2_digit_pertama):
        """Mengembalikan nama sheet berdasarkan 2 digit pertama kode"""
        sheet_mapping = {
            '01': 'CIGUGUR',
            '02': 'MELONG', 
            '03': 'LG '
        }
        return sheet_mapping.get(kode_2_digit_pertama, None)
    
    def parse_kode_8_digit_extract(kode):
        """Memecah kode 8 digit menjadi 4 bagian: 2-2-2-2"""
        if len(kode) != 8:
            return None, None, None, None
        
        digit_1_2 = kode[:2]   # Menentukan sheet
        digit_3_4 = kode[2:4]  # Kolom B
        digit_5_6 = kode[4:6]  # Kolom C  
        digit_7_8 = kode[6:8]  # Kolom D
        
        return digit_1_2, digit_3_4, digit_5_6, digit_7_8
    
    def find_row_by_code_extract(worksheet, digit_3_4, digit_5_6, digit_7_8):
        """Mencari row yang sesuai berdasarkan kode di kolom B, C, D"""
        for row in range(2, worksheet.max_row + 1):  # Mulai dari row 2 (skip header)
            col_b_value = str(worksheet[f'B{row}'].value or '').zfill(2)
            col_c_value = str(worksheet[f'C{row}'].value or '').zfill(2)
            col_d_value = str(worksheet[f'D{row}'].value or '').zfill(2)
            
            if (col_b_value == digit_3_4 and 
                col_c_value == digit_5_6 and 
                col_d_value == digit_7_8):
                return row
        
        return None
    
    def format_date_only(date_value):
        """Mengkonversi nilai tanggal dari Excel menjadi string tanggal saja (tanpa waktu 00:00:00)"""
        if not date_value or date_value == '':
            return ''
        
        # Jika sudah berupa datetime object
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d')
        
        # Jika berupa string yang mengandung datetime
        if isinstance(date_value, str):
            try:
                # Coba parse berbagai format tanggal
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                    try:
                        parsed_date = datetime.strptime(date_value, fmt)
                        return parsed_date.strftime('%Y-%m-%d')
                    except:
                        continue
            except:
                pass
        
        # Jika gagal parsing, return as string
        return str(date_value)
    
    # OPTIMIZATION: Load workbooks once outside the loop
    workbooks = {}
    worksheets_cache = {}
    
    # Load all workbooks at the beginning
    for year, excel_file in master_files.items():
        try:
            workbooks[year] = openpyxl.load_workbook(excel_file, data_only=True)
            # Cache all worksheets for this workbook
            worksheets_cache[year] = {}
            for sheet_name in ['CIGUGUR', 'MELONG', 'LG ']:
                if sheet_name in workbooks[year].sheetnames:
                    worksheets_cache[year][sheet_name] = workbooks[year][sheet_name]
        except Exception as e:
            st.warning(f"Warning: Tidak dapat memuat file Excel untuk tahun {year}: {str(e)}")
    
    def extract_data_excel_optimized(row):
        kode = str(row['Kode_8_Digit'])
        bulan = row['Bulan']
        tahun = str(row['Tahun'])
        
        # Validasi data
        if len(kode) != 8:
            return pd.Series([None] * len(extract_cols), index=extract_cols)
        
        # Cek apakah workbook untuk tahun ini sudah dimuat
        if tahun not in workbooks:
            return pd.Series([None] * len(extract_cols), index=extract_cols)
        
        # Parse kode 8 digit
        digit_1_2, digit_3_4, digit_5_6, digit_7_8 = parse_kode_8_digit_extract(kode)
        
        # Tentukan sheet
        sheet_name = get_sheet_name_extract(digit_1_2)
        if not sheet_name:
            return pd.Series([None] * len(extract_cols), index=extract_cols)
        
        # Dapatkan kolom berdasarkan bulan
        month_cols = get_month_columns_extract(bulan)
        if not month_cols:
            return pd.Series([None] * len(extract_cols), index=extract_cols)
        
        try:
            # Use cached worksheet instead of loading file again
            if tahun not in worksheets_cache or sheet_name not in worksheets_cache[tahun]:
                return pd.Series([None] * len(extract_cols), index=extract_cols)
            
            worksheet = worksheets_cache[tahun][sheet_name]
            
            # Cari row berdasarkan kode di kolom B, C, D
            target_row = find_row_by_code_extract(worksheet, digit_3_4, digit_5_6, digit_7_8)
            if not target_row:
                return pd.Series([None] * len(extract_cols), index=extract_cols)
            
            # Extract data dari 4 kolom
            nama_penghuni = worksheet[f'{month_cols[0]}{target_row}'].value or ''
            tanggal_perjanjian_raw = worksheet[f'{month_cols[1]}{target_row}'].value or ''
            sewa_hunian = worksheet[f'{month_cols[2]}{target_row}'].value or ''
            sewa_lahan = worksheet[f'{month_cols[3]}{target_row}'].value or ''
            
            # Format tanggal tanpa waktu
            tanggal_perjanjian = format_date_only(tanggal_perjanjian_raw)
            
            extracted_data = [
                str(nama_penghuni),
                tanggal_perjanjian,
                str(sewa_hunian),
                str(sewa_lahan)
            ]
            
            return pd.Series(extracted_data, index=extract_cols)
                
        except Exception:
            return pd.Series([None] * len(extract_cols), index=extract_cols)
    
    # Terapkan ekstraksi dengan fungsi yang dioptimasi
    extraction_results = df_setortunai.apply(extract_data_excel_optimized, axis=1)
    
    # Gabungkan dengan dataframe asli
    df_with_extract = pd.concat([df_setortunai, extraction_results], axis=1)
    
    return df_with_extract

def calculate_denda(df_with_extract):
    """Kalkulasi denda berdasarkan selisih Credit Transaction dengan total sewa"""
    def parse_kode_for_mapping(kode):
        if len(str(kode)) >= 8:
            rusunawa_code = str(kode)[:2]
            gedung_code = str(kode)[2:4]
            lantai_code = str(kode)[4:6]
            hunian_code = str(kode)[6:8]
            
            rusunawa_mapping = {
                '01': 'Cigugur Tengah',
                '02': 'Cibeureum',
                '03': 'Leuwigajah'
            }
            
            gedung_mapping = {
                '01': 'A', '02': 'B', '03': 'C', '04': 'D'
            }
            
            lantai_mapping = {
                '01': 'I', '02': 'II', '03': 'III', '04': 'IV', '05': 'V'
            }
            
            return {
                'Rusunawa': rusunawa_mapping.get(rusunawa_code, ''),
                'Gedung': gedung_mapping.get(gedung_code, ''),
                'Lantai': lantai_mapping.get(lantai_code, ''),
                'No Hunian': int(hunian_code) if hunian_code.isdigit() else 0
            }
        return {'Rusunawa': '', 'Gedung': '', 'Lantai': '', 'No Hunian': 0}
    
    def calculate_denda_amount(row):
        try:
            # Ambil nilai Credit Transaction dari bank statement
            credit_transaction = to_numeric_safe(row.get('Credit Transaction', 0))
            
            # Ambil nilai sewa dari Master Excel
            sewa_hunian = to_numeric_safe(row.get('Sewa Hunian', 0))
            sewa_lahan = to_numeric_safe(row.get('Sewa Lahan Lantai 1', 0))
            
            # Total sewa yang seharusnya dibayar
            total_sewa = sewa_hunian + sewa_lahan
            
            # Denda = Credit Transaction - Total Sewa
            # Jika Credit Transaction > Total Sewa, maka ada denda
            denda = credit_transaction - total_sewa
            
            # Denda hanya dihitung jika positif (ada kelebihan pembayaran yang menunjukkan denda)
            return max(0, round(denda, 0))
        except:
            return 0.0
    
    # Tambahkan kolom mapping
    mapping_results = df_with_extract['Kode_8_Digit'].apply(parse_kode_for_mapping)
    for key in ['Rusunawa', 'Gedung', 'Lantai', 'No Hunian']:
        df_with_extract[key] = [result[key] for result in mapping_results]
    
    # Kalkulasi denda berdasarkan selisih Credit Transaction dengan total sewa
    df_with_extract['Denda'] = df_with_extract.apply(calculate_denda_amount, axis=1)
    
    return df_with_extract

def input_to_excel_master(df_final, master_files):
    """Input data ke Excel Master dengan backup dan pewarnaan (disesuaikan dengan code baru) - OPTIMIZED VERSION"""
    import shutil
    import os
    from openpyxl.styles import PatternFill
    import openpyxl
    
    # === DEFINISI FUNGSI INTERNAL ===
    
    def get_month_columns(bulan):
        """Mengembalikan kolom-kolom target berdasarkan bulan."""
        month_mapping = {
            'Januari': {'posting_date_cols': ['P', 'Q'], 'denda_col': 'M', 'tambahan_col': 'R', 'color_range': ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']},
            'Februari': {'posting_date_cols': ['Z', 'AA'], 'denda_col': 'W', 'tambahan_col': 'AB', 'color_range': ['S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']},
            'Maret': {'posting_date_cols': ['AJ', 'AK'], 'denda_col': 'AG', 'tambahan_col': 'AL', 'color_range': ['AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL']},
            'April': {'posting_date_cols': ['AT', 'AU'], 'denda_col': 'AQ', 'tambahan_col': 'AV', 'color_range': ['AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV']},
            'Mei': {'posting_date_cols': ['BD', 'BE'], 'denda_col': 'BA', 'tambahan_col': 'BF', 'color_range': ['AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF']},
            'Juni': {'posting_date_cols': ['BN', 'BO'], 'denda_col': 'BK', 'tambahan_col': 'BP', 'color_range': ['BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP']},
            'Juli': {'posting_date_cols': ['BX', 'BY'], 'denda_col': 'BU', 'tambahan_col': 'BZ', 'color_range': ['BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ']},
            'Agustus': {'posting_date_cols': ['CH', 'CI'], 'denda_col': 'CE', 'tambahan_col': 'CJ', 'color_range': ['CA', 'CB', 'CC', 'CD', 'CE', 'CF', 'CG', 'CH', 'CI', 'CJ']},
            'September': {'posting_date_cols': ['CR', 'CS'], 'denda_col': 'CO', 'tambahan_col': 'CT', 'color_range': ['CK', 'CL', 'CM', 'CN', 'CO', 'CP', 'CQ', 'CR', 'CS', 'CT']},
            'Oktober': {'posting_date_cols': ['DB', 'DC'], 'denda_col': 'CY', 'tambahan_col': 'DD', 'color_range': ['CU', 'CV', 'CW', 'CX', 'CY', 'CZ', 'DA', 'DB', 'DC', 'DD']},
            'November': {'posting_date_cols': ['DL', 'DM'], 'denda_col': 'DI', 'tambahan_col': 'DN', 'color_range': ['DE', 'DF', 'DG', 'DH', 'DI', 'DJ', 'DK', 'DL', 'DM', 'DN']},
            'Desember': {'posting_date_cols': ['DV', 'DW'], 'denda_col': 'DS', 'tambahan_col': 'DX', 'color_range': ['DO', 'DP', 'DQ', 'DR', 'DS', 'DT', 'DU', 'DV', 'DW', 'DX']}
        }
        return month_mapping.get(bulan)

    def get_sheet_name(kode_2_digit):
        """Mengembalikan nama sheet berdasarkan 2 digit pertama kode."""
        return {'01': 'CIGUGUR', '02': 'MELONG', '03': 'LG '}.get(kode_2_digit)

    def parse_kode_8_digit(kode):
        """Memecah kode 8 digit menjadi 4 bagian."""
        return (kode[:2], kode[2:4], kode[4:6], kode[6:8]) if len(kode) == 8 else (None, None, None, None)

    def find_row_by_code(worksheet, d3_4, d5_6, d7_8):
        """Mencari baris yang sesuai berdasarkan kode di kolom B, C, D."""
        for row in range(2, worksheet.max_row + 1):
            if (str(worksheet[f'B{row}'].value or '').zfill(2) == d3_4 and
                str(worksheet[f'C{row}'].value or '').zfill(2) == d5_6 and
                str(worksheet[f'D{row}'].value or '').zfill(2) == d7_8):
                return row
        return None

    def is_cell_filled(worksheet, cell_address):
        """Cek apakah sel sudah terisi data."""
        cell_value = worksheet[cell_address].value
        return cell_value is not None and str(cell_value).strip() != ''

    def create_backup_file(original_file):
        """Membuat backup file dengan timestamp."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name, ext = os.path.splitext(original_file)
        backup_path = f"{name}_BACKUP_{timestamp}{ext}"
        shutil.copy2(original_file, backup_path)
        return backup_path

    def convert_date_format(date_str):
        """Mengkonversi format tanggal string ke objek datetime."""
        try:
            if pd.isna(date_str) or date_str == '' or date_str == 'nan':
                return None
            
            # Coba parse format dd/mm/yyyy atau mm/dd/yyyy
            return pd.to_datetime(str(date_str)).to_pydatetime()
        except (ValueError, TypeError):
            return None # Kembalikan None jika parsing gagal

    def input_data_to_excel_v2_silent_optimized(df_data):
        """Versi yang dioptimasi: Input data ke Excel dengan minimal I/O operations."""
        results = {'success': 0, 'skipped': 0, 'failed': 0, 'errors': [], 'skipped_details': [], 'success_details': [], 'failed_details': [], 'backup_files': {}}
        
        # Group data by year for batch processing
        for tahun, group_data in df_data.groupby('Tahun'):
            if not tahun: 
                continue
                
            # Cari file master berdasarkan tahun
            excel_file = None
            if str(tahun) == '2024' and '2024' in master_files:
                excel_file = master_files['2024']
            elif str(tahun) == '2025' and '2025' in master_files:
                excel_file = master_files['2025']
            
            if not excel_file or not os.path.exists(excel_file):
                results['errors'].append(f"File tidak ditemukan untuk tahun {tahun}")
                continue
            
            try:
                # OPTIMIZATION: Create backup and load workbook ONCE per year
                backup_path = create_backup_file(excel_file)
                results['backup_files'][str(tahun)] = backup_path
                workbook = openpyxl.load_workbook(backup_path)
                
                # Cache all worksheets to avoid repeated sheet access
                worksheets_cache = {}
                for sheet_name in ['CIGUGUR', 'MELONG', 'LG ']:
                    if sheet_name in workbook.sheetnames:
                        worksheets_cache[sheet_name] = workbook[sheet_name]
                
                # Process all rows for this year in memory
                for _, row in group_data.iterrows():
                    kode, bulan, p_date, denda = str(row['Kode_8_Digit']), row['Bulan'], row['Posting Date'], row['Denda']
                    
                    # Konversi format tanggal ke objek datetime
                    p_date_converted = convert_date_format(p_date)
                    
                    d1, d2, d3, d4 = parse_kode_8_digit(kode)
                    sheet_name = get_sheet_name(d1) if d1 else None
                    month_cols = get_month_columns(bulan)

                    if not all([d1, sheet_name, month_cols]):
                        results['failed'] += 1
                        results['failed_details'].append({'Kode_8_Digit': kode, 'Bulan': bulan, 'Reason': 'Kode/Sheet/Bulan tidak valid'})
                        continue

                    if sheet_name not in worksheets_cache:
                        results['failed'] += 1
                        results['failed_details'].append({'Kode_8_Digit': kode, 'Bulan': bulan, 'Reason': f"Sheet '{sheet_name}' tidak ada"})
                        continue

                    # Use cached worksheet
                    worksheet = worksheets_cache[sheet_name]
                    target_row = find_row_by_code(worksheet, d2, d3, d4)
                    if not target_row:
                        results['failed'] += 1
                        results['failed_details'].append({'Kode_8_Digit': kode, 'Bulan': bulan, 'Reason': f'Kombinasi kode tidak ditemukan'})
                        continue
                    
                    pdc1, pdc2, dc, tc = month_cols['posting_date_cols'][0], month_cols['posting_date_cols'][1], month_cols['denda_col'], month_cols['tambahan_col']
                    c1_addr, c2_addr, c3_addr, c4_addr = f'{pdc1}{target_row}', f'{pdc2}{target_row}', f'{dc}{target_row}', f'{tc}{target_row}'
                    
                    if is_cell_filled(worksheet, c1_addr) or is_cell_filled(worksheet, c2_addr) or is_cell_filled(worksheet, c3_addr):
                        results['skipped'] += 1
                        results['skipped_details'].append({'Kode_8_Digit': kode, 'Bulan': bulan, 'Reason': f"Sudah terisi di {sheet_name} baris {target_row}"})
                        continue
                    
                    try:
                        # Perform all changes in memory
                        
                        # === PERUBAHAN DI SINI ===
                        if p_date_converted is not None:
                            # Tulis objek datetime langsung ke sel
                            worksheet[c1_addr] = p_date_converted
                            worksheet[c2_addr] = p_date_converted
                            
                            # Terapkan format tanggal DD-MM-YYYY
                            worksheet[c1_addr].number_format = 'DD/MM/YYYY'
                            worksheet[c2_addr].number_format = 'DD/MM/YYYY'
                        else:
                            # Jika tanggal tidak valid, tulis string aslinya
                            worksheet[c1_addr] = p_date 
                            worksheet[c2_addr] = p_date
                        # === AKHIR PERUBAHAN ===
                        
                        worksheet[c3_addr] = denda
                        worksheet[c4_addr] = 1  # Tambahan angka 1
                        
                        # Apply formatting in memory (pewarnaan)
                        from openpyxl.styles import PatternFill
                        fill_color = PatternFill(start_color='D8E4BC', end_color='D8E4BC', fill_type='solid')
                        for col in month_cols['color_range']:
                            worksheet[f'{col}{target_row}'].fill = fill_color
                        
                        results['success'] += 1
                        results['success_details'].append({'Kode_8_Digit': kode, 'Bulan': bulan, 'Sheet': sheet_name, 'Row': target_row, 'Denda': denda})
                    except Exception as e:
                        results['failed'] += 1
                        results['failed_details'].append({'Kode_8_Digit': kode, 'Bulan': bulan, 'Reason': f'Error input: {str(e)}'})

                # Hide columns A, B, C, D in all relevant sheets before saving
                for sheet_name in ['CIGUGUR', 'MELONG', 'LG ']:
                    if sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        # Hide columns A, B, C, D
                        ws.column_dimensions['A'].hidden = True
                        ws.column_dimensions['B'].hidden = True
                        ws.column_dimensions['C'].hidden = True
                        ws.column_dimensions['D'].hidden = True
                
                # OPTIMIZATION: Save workbook ONCE after all changes are done
                workbook.save(backup_path)
                
            except Exception as e:
                results['errors'].append(f"Error proses file tahun {tahun}: {str(e)}")
                
        return results

    # === EKSEKUSI UTAMA ===
    
    # Filter data valid secara langsung
    valid_data = df_final[
        (df_final['Kode_8_Digit'].str.len() == 8) & 
        (df_final['Bulan'].notna()) & (df_final['Bulan'] != '') &
        (df_final['Tahun'].notna()) & (df_final['Tahun'] != '') &
        (df_final['Posting Date'].notna()) &
        (df_final['Denda'].notna())
    ].copy()
    
    if not valid_data.empty:
        # Jalankan fungsi yang sudah dioptimasi
        results = input_data_to_excel_v2_silent_optimized(valid_data)
    else:
        results = {'success': 0, 'skipped': 0, 'failed': 0, 'errors': ['Tidak ada data valid untuk diinput setelah difilter.'], 'skipped_details': [], 'success_details': [], 'failed_details': [], 'backup_files': {}}
    
    return df_final, results

def create_export_excel(results, valid_data, df_final, df_non_rusun):
    """Membuat file Excel hasil dengan beberapa sheet"""
    # Buat dataframe untuk export dengan kolom status
    df_export_status = valid_data.copy()
    df_export_status['Status_Input'] = 'Belum Diproses'
    df_export_status['Keterangan_Input'] = ''
    df_export_status['Nilai_Denda_Input'] = ''
    
    # Update status berdasarkan hasil
    for success_detail in results.get('success_details', []):
        kode = success_detail['Kode_8_Digit']
        bulan = success_detail['Bulan']
        mask = (df_export_status['Kode_8_Digit'] == kode) & (df_export_status['Bulan'] == bulan)
        df_export_status.loc[mask, 'Status_Input'] = 'Berhasil Input'
        df_export_status.loc[mask, 'Keterangan_Input'] = f"Data berhasil diinput ke {success_detail['Sheet']} row {success_detail['Row']}"
        df_export_status.loc[mask, 'Nilai_Denda_Input'] = success_detail['Denda']
    
    for skip_detail in results.get('skipped_details', []):
        kode = skip_detail['Kode_8_Digit']
        bulan = skip_detail['Bulan']
        mask = (df_export_status['Kode_8_Digit'] == kode) & (df_export_status['Bulan'] == bulan)
        df_export_status.loc[mask, 'Status_Input'] = 'Skip - Sudah Terisi'
        df_export_status.loc[mask, 'Keterangan_Input'] = skip_detail['Reason']
    
    for failed_detail in results.get('failed_details', []):
        kode = failed_detail['Kode_8_Digit']
        bulan = failed_detail['Bulan']
        mask = (df_export_status['Kode_8_Digit'] == kode) & (df_export_status['Bulan'] == bulan)
        df_export_status.loc[mask, 'Status_Input'] = 'Gagal'
        df_export_status.loc[mask, 'Keterangan_Input'] = failed_detail['Reason']
    
    # Fungsi format
    def format_date_ddmmmyy(date_str):
        try:
            if pd.isna(date_str) or date_str == '':
                return ''
            date_obj = pd.to_datetime(date_str, format='%d/%m/%Y')
            return date_obj.strftime('%d-%b-%y')
        except:
            return str(date_str) if date_str else ''
    
    def format_month_mmmyy(bulan, tahun):
        try:
            if pd.isna(bulan) or bulan == '' or pd.isna(tahun) or tahun == '':
                return ''
            
            month_mapping = {
                'Januari': 'Jan', 'Februari': 'Feb', 'Maret': 'Mar',
                'April': 'Apr', 'Mei': 'May', 'Juni': 'Jun',
                'Juli': 'Jul', 'Agustus': 'Aug', 'September': 'Sep',
                'Oktober': 'Oct', 'November': 'Nov', 'Desember': 'Dec'
            }
            
            month_abbr = month_mapping.get(bulan, bulan)
            year_short = str(tahun)[-2:]
            return f"{month_abbr}-{year_short}"
        except:
            return f"{bulan}-{str(tahun)[-2:]}" if bulan and tahun else ''
    
    def format_gedung_lantai_hunian(gedung, lantai, no_hunian):
        """Memformat kolom gedung, lantai, dan no hunian secara terpisah"""
        try:
            # Format Gedung
            gedung_formatted = str(gedung) if pd.notna(gedung) else ''
            
            # Format Lantai dengan angka romawi
            lantai_romawi = {'1': 'I', '2': 'II', '3': 'III', '4': 'IV', '5': 'V', 'I': 'I', 'II': 'II', 'III': 'III', 'IV': 'IV', 'V': 'V'}
            lantai_formatted = lantai_romawi.get(str(lantai), str(lantai)) if pd.notna(lantai) else ''
            
            # Format No Hunian dengan leading zero
            no_hunian_formatted = str(no_hunian).zfill(2) if pd.notna(no_hunian) and str(no_hunian).isdigit() else str(no_hunian) if pd.notna(no_hunian) else ''
            
            return gedung_formatted, lantai_formatted, no_hunian_formatted
        except:
            return '', '', ''
    
    # BARU: Fungsi untuk menyesuaikan lebar kolom
    def auto_adjust_column_width(worksheet):
        """Menyesuaikan lebar kolom di worksheet berdasarkan konten terpanjang."""
        for col in worksheet.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Siapkan data untuk sheet Kasda dan Denda
    df_kasda = df_final.copy()
    df_kasda = df_kasda.reset_index(drop=True)
    df_kasda['No Urut'] = range(1, len(df_kasda) + 1)
    
    # Format kolom Gedung, Lantai, No Hunian secara terpisah untuk sheet Kasda
    gedung_lantai_hunian = df_kasda.apply(
        lambda x: format_gedung_lantai_hunian(x.get('Gedung'), x.get('Lantai'), x.get('No Hunian')), 
        axis=1
    )
    
    # Format kolom untuk sheet Kasda
    df_kasda_export = pd.DataFrame({
        'No Urut': df_kasda['No Urut'],
        'Tanggal Setor': df_kasda['Posting Date'],
        'Tanggal Kasda': df_kasda['Posting Date'],
        'Rusunawa': df_kasda.get('Rusunawa', ''),
        'Gedung': [item[0] for item in gedung_lantai_hunian],
        'Lantai': [item[1] for item in gedung_lantai_hunian],
        'No Hunian': [item[2] for item in gedung_lantai_hunian],
        'Nama Penghuni': df_kasda.get('Nama Penghuni', ''),
        'Tanggal Perjanjian': df_kasda.get('Tanggal Perjanjian Sewa', '').apply(format_date_ddmmmyy),
        'Sewa Hunian': df_kasda.get('Sewa Hunian', 0).apply(to_numeric_safe),
        'Sewa Lantai 1': df_kasda.get('Sewa Lahan Lantai 1', 0).apply(to_numeric_safe),
        'Denda': None,
        'Jumlah': (df_kasda.get('Sewa Hunian', 0).apply(to_numeric_safe) + 
                  df_kasda.get('Sewa Lahan Lantai 1', 0).apply(to_numeric_safe)),
        'Bulan': df_kasda.apply(lambda x: convert_to_first_of_month_date(x.get('Bulan'), x.get('Tahun')), axis=1)
    })
    
    # Data untuk sheet Denda
    df_denda_filter = df_final[df_final.get('Denda', 0).apply(to_numeric_safe) > 0].copy()
    df_denda = df_denda_filter.reset_index(drop=True)
    df_denda['No Urut'] = range(1, len(df_denda) + 1)
    
    # Format kolom Gedung, Lantai, No Hunian secara terpisah untuk sheet Denda
    gedung_lantai_hunian_denda = df_denda.apply(
        lambda x: format_gedung_lantai_hunian(x.get('Gedung'), x.get('Lantai'), x.get('No Hunian')), 
        axis=1
    )
    
    df_denda_export = pd.DataFrame({
        'No Urut': df_denda['No Urut'],
        'Tanggal Setor': df_denda['Posting Date'],
        'Tanggal Kasda': df_denda['Posting Date'],
        'Rusunawa': df_denda.get('Rusunawa', ''),
        'Gedung': [item[0] for item in gedung_lantai_hunian_denda],
        'Lantai': [item[1] for item in gedung_lantai_hunian_denda],
        'No Hunian': [item[2] for item in gedung_lantai_hunian_denda],
        'Nama Penghuni': df_denda.get('Nama Penghuni', ''),
        'Tanggal Perjanjian': df_denda.get('Tanggal Perjanjian Sewa', '').apply(format_date_ddmmmyy),
        'Denda': df_denda.get('Denda', 0).apply(to_numeric_safe),
        'Jumlah': df_denda.get('Denda', 0).apply(to_numeric_safe),
        'Bulan': df_denda.apply(lambda x: convert_to_first_of_month_date(x.get('Bulan'), x.get('Tahun')), axis=1)
    })
    
    # Konversi kolom numeric di df_export_status
    money_columns_status = ['Credit Transaction', 'Balance', 'Sewa Hunian', 'Sewa Lahan Lantai 1', 'Denda']
    for col in money_columns_status:
        if col in df_export_status.columns:
            df_export_status[col] = df_export_status[col].apply(to_numeric_safe)
    
    # Export ke file Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_dir = os.path.join(tempfile.gettempdir(), "rusun_app", "output")
    os.makedirs(temp_dir, exist_ok=True)
    export_filename = os.path.join(temp_dir, f"Status_Input_Excel_{timestamp}.xlsx")
    
    with pd.ExcelWriter(export_filename, engine='openpyxl') as writer:
        # Sheet 1: Data status input
        df_export_status.to_excel(writer, sheet_name='Status_Input', index=False)
        
        # Sheet 2: Data non-rusun untuk cek manual
        if len(df_non_rusun) > 0:
            df_non_rusun_export = df_non_rusun.copy()
            for col in money_columns_status:
                if col in df_non_rusun_export.columns:
                    df_non_rusun_export[col] = df_non_rusun_export[col].apply(to_numeric_safe)
            df_non_rusun_export.to_excel(writer, sheet_name='Cek Manual', index=False)
        
        # Sheet 3: Data Kasda
        df_kasda_export.to_excel(writer, sheet_name='Kasda', index=False)
        
        # Sheet 4: Data Denda
        if len(df_denda_export) > 0:
            df_denda_export.to_excel(writer, sheet_name='Denda', index=False)
        
        # Format Excel dengan numeric format
        workbook = writer.book
        
        # 1. Definisikan format yang diinginkan
        money_format = '#,##0'
        date_format = 'DD-MM-YYYY'
        month_year_format = 'MMM-YY'  # BARU: Format tampilan Bulan-Tahun

        # 2. Definisikan kolom-kolom tanggal
        date_cols = ['Tanggal Setor', 'Tanggal Kasda', 'Tanggal Perjanjian']
        
        # 3. Definisikan kolom-kolom uang per sheet
        money_cols = {
            'Status_Input': ['Credit Transaction', 'Balance', 'Sewa Hunian', 'Sewa Lahan Lantai 1', 'Denda'],
            'Cek Manual': money_columns_status,
            'Kasda': ['Sewa Hunian', 'Sewa Lantai 1', 'Jumlah'],
            'Denda': ['Denda', 'Jumlah']
        }
        
        # 4. Mapping sheet ke dataframe
        sheets_to_write = {
            'Status_Input': df_export_status,
            'Cek Manual': df_non_rusun if len(df_non_rusun) > 0 else pd.DataFrame(),
            'Kasda': df_kasda_export,
            'Denda': df_denda_export if len(df_denda_export) > 0 else pd.DataFrame()
        }
        
        # Loop melalui setiap sheet untuk menerapkan format
        for sheet_name, df_to_format in sheets_to_write.items():
            if sheet_name in workbook.sheetnames and not df_to_format.empty:
                worksheet = workbook[sheet_name]
                
                # Terapkan format angka untuk kolom uang
                sheet_money_cols = money_cols.get(sheet_name, [])
                for col_name in sheet_money_cols:
                    if col_name in df_to_format.columns:
                        col_idx = df_to_format.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        
                        for row in range(2, len(df_to_format) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = money_format
                
                # Terapkan format tanggal untuk kolom tanggal
                for col_name in date_cols:
                    if col_name in df_to_format.columns:
                        col_idx = df_to_format.columns.get_loc(col_name) + 1
                        col_letter = get_column_letter(col_idx)
                        
                        for row in range(2, len(df_to_format) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                cell.number_format = date_format

                # --- PERUBAHAN DI SINI ---
                # Terapkan format mmm-yy untuk kolom 'Bulan' di sheet Kasda dan Denda
                if sheet_name in ['Kasda', 'Denda']:
                    if 'Bulan' in df_to_format.columns:
                        col_idx = df_to_format.columns.get_loc('Bulan') + 1
                        col_letter = get_column_letter(col_idx)
                        
                        for row in range(2, len(df_to_format) + 2):
                            cell = worksheet[f'{col_letter}{row}']
                            if cell.value is not None:
                                # Terapkan format tampilan, bukan format teks
                                cell.number_format = month_year_format
                # --- AKHIR PERUBAHAN ---

                # Terapkan penyesuaian lebar kolom otomatis
                auto_adjust_column_width(worksheet)
    
    return export_filename

# Main app
def main():
    init_session_state()
    
    # Header
    st.title("🏢 Aplikasi Penerimaan Rusun")
    st.markdown("### Upload file dan proses otomatis dengan satu klik")
    st.markdown("---")
    
    # Cek apakah sudah ada hasil yang bisa didownload
    if hasattr(st.session_state, 'export_file') and st.session_state.export_file:
        # Tampilkan area download hasil
        st.success("🎉 **Proses selesai!** File hasil sudah siap untuk didownload.")
        
        col1, col2 = st.columns(2)
        
        # Download File Excel Status/Report
        with col1:
            st.subheader("📊 File Laporan Status")
            
            if os.path.exists(st.session_state.export_file):
                file_size = os.path.getsize(st.session_state.export_file)
                filename = os.path.basename(st.session_state.export_file)
                
                st.info(f"📁 **File**: {filename}")
                st.info(f"📊 **Ukuran**: {file_size:,} bytes")
                
                with open(st.session_state.export_file, "rb") as file:
                    file_data = file.read()
                    
                    st.download_button(
                        label="📥 Download Laporan Status",
                        data=file_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
            else:
                st.error("❌ File export tidak ditemukan")
        
        # Download File Master Excel yang sudah diupdate
        with col2:
            st.subheader("📋 File Master Excel Terupdate")
            
            if hasattr(st.session_state, 'results') and 'backup_files' in st.session_state.results:
                backup_files = st.session_state.results['backup_files']
                
                if backup_files:
                    for tahun, backup_path in backup_files.items():
                        if os.path.exists(backup_path):
                            file_size = os.path.getsize(backup_path)
                            filename = os.path.basename(backup_path)
                            
                            st.info(f"📁 **Master {tahun}**: {filename}")
                            st.info(f"📊 **Ukuran**: {file_size:,} bytes")
                            
                            with open(backup_path, "rb") as file:
                                file_data = file.read()
                                
                                st.download_button(
                                    label=f"📥 Download Master {tahun}",
                                    data=file_data,
                                    file_name=f"Master_Updated_{tahun}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    type="secondary",
                                    use_container_width=True,
                                    key=f"download_master_{tahun}"
                                )
                else:
                    st.warning("⚠️ Tidak ada file Master yang diupdate")
            else:
                st.warning("⚠️ Data hasil tidak tersedia")
        
        # Ringkasan hasil
        if hasattr(st.session_state, 'results') and st.session_state.results:
            st.markdown("---")
            st.subheader("📊 Ringkasan Hasil Proses")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("✅ Berhasil Input", st.session_state.results['success'])
            with col2:
                st.metric("⏭️ Dilewati", st.session_state.results['skipped'])
            with col3:
                st.metric("❌ Gagal", st.session_state.results['failed'])
        
        # Tombol untuk mulai baru
        st.markdown("---")
        if st.button("🔄 Proses File Baru", type="secondary", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    else:
        # Area upload file dan proses otomatis
        st.subheader("📤 Upload File untuk Diproses")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown("#### 📄 PDF Bank Statement")
            bank_file = st.file_uploader(
                "Pilih file PDF Bank Statement", 
                type=["pdf"], 
                key="bank_file",
                help="Upload file PDF yang berisi data transaksi bank"
            )
            
            if bank_file:
                st.success(f"✅ File terpilih: {bank_file.name}")
                st.info(f"📊 Ukuran: {bank_file.size:,} bytes")
        
        with col2:
            st.markdown("#### 📊 Master Excel")
            master_2024 = st.file_uploader(
                "Master Excel 2024", 
                type=["xlsx"], 
                key="master_2024",
                help="File Excel master untuk data tahun 2024"
            )
            
            master_2025 = st.file_uploader(
                "Master Excel 2025", 
                type=["xlsx"], 
                key="master_2025",
                help="File Excel master untuk data tahun 2025"
            )
            
            if master_2024:
                st.success(f"✅ 2024: {master_2024.name}")
            if master_2025:
                st.success(f"✅ 2025: {master_2025.name}")
        
        st.markdown("---")
        
        # Tombol proses otomatis
        if bank_file:
            if st.button("🚀 **PROSES OTOMATIS - SEMUA LANGKAH**", type="primary", use_container_width=True):
                
                # Progress tracking
                progress_container = st.container()
                status_container = st.container()
                
                with progress_container:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                
                try:
                    # Step 1: Simpan file ke temporary
                    with status_container:
                        status_text.info("🔄 **Langkah 1/6:** Menyimpan file yang diupload...")
                    progress_bar.progress(10)
                    
                    temp_pdf = save_uploadedfile_temp(bank_file, "bank_files")
                    
                    master_files = {}
                    if master_2024:
                        temp_master_2024 = save_uploadedfile_temp(master_2024, "Master Data")
                        master_files['2024'] = temp_master_2024
                    if master_2025:
                        temp_master_2025 = save_uploadedfile_temp(master_2025, "Master Data")
                        master_files['2025'] = temp_master_2025
                    
                    # Step 2: Proses PDF
                    status_text.info("🔄 **Langkah 2/6:** Mengekstrak data dari PDF Bank Statement...")
                    progress_bar.progress(20)
                    
                    df_bank = process_pdf(temp_pdf)
                    
                    if len(df_bank) == 0:
                        st.error("❌ Tidak ada data yang bisa diekstrak dari PDF")
                        return
                    
                    # Step 3: Ekstrak SETORTUNAI
                    status_text.info(f"🔄 **Langkah 3/6:** Mengekstrak data SETORTUNAI dari {len(df_bank)} transaksi...")
                    progress_bar.progress(35)
                    
                    df_setortunai, df_non_rusun = process_setortunai(df_bank)
                    
                    # Step 4: Filter data dan ekstrak dari Master
                    status_text.info("🔄 **Langkah 4/6:** Memfilter data dan mengekstrak dari Master Excel...")
                    progress_bar.progress(50)
                    
                    df_setortunai_filtered, df_non_rusun_new = filter_incomplete_data(df_setortunai, df_non_rusun)
                    
                    if master_files:
                        df_with_extract = extract_from_master_excel(df_setortunai_filtered, master_files)
                    else:
                        df_with_extract = df_setortunai_filtered
                        df_with_extract['Nama Penghuni'] = None
                        df_with_extract['Tanggal Perjanjian Sewa'] = None
                        df_with_extract['Sewa Hunian'] = 0
                        df_with_extract['Sewa Lahan Lantai 1'] = 0
                    
                    df_final = calculate_denda(df_with_extract)
                    
                    # Step 5: Input ke Excel Master
                    status_text.info("🔄 **Langkah 5/6:** Memasukkan data ke Excel Master...")
                    progress_bar.progress(70)
                    
                    if master_files:
                        valid_data, results = input_to_excel_master(df_final, master_files)
                    else:
                        # Jika tidak ada master files, buat hasil dummy
                        valid_data = df_final
                        results = {
                            'success': 0, 'skipped': 0, 'failed': 0, 
                            'errors': ['Tidak ada file Master Excel untuk input'], 
                            'skipped_details': [], 'success_details': [], 'failed_details': [], 
                            'backup_files': {}
                        }
                    
                    # Step 6: Export hasil
                    status_text.info("🔄 **Langkah 6/6:** Membuat file export hasil...")
                    progress_bar.progress(85)
                    
                    export_file = create_export_excel(results, valid_data, df_final, df_non_rusun_new)
                    
                    # Simpan hasil ke session state
                    st.session_state.df_bank = df_bank
                    st.session_state.df_setortunai = df_setortunai
                    st.session_state.df_non_rusun = df_non_rusun_new
                    st.session_state.df_final = df_final
                    st.session_state.valid_data = valid_data
                    st.session_state.results = results
                    st.session_state.export_file = export_file
                    st.session_state.master_files = master_files
                    
                    # Selesai
                    progress_bar.progress(100)
                    status_text.success("✅ **Proses selesai!** Semua langkah berhasil dijalankan.")
                    
                    # Tampilkan ringkasan
                    st.success("🎉 **Proses Otomatis Berhasil Diselesaikan!**")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("📊 Total Transaksi", len(df_bank))
                    with col2:
                        st.metric("🏢 Data SETORTUNAI", len(df_setortunai))
                    with col3:
                        st.metric("📄 Data NON-RUSUN", len(df_non_rusun_new))
                    with col4:
                        denda_count = (df_final['Denda'] > 0).sum()
                        st.metric("💰 Data dengan Denda", denda_count)
                    
                    if master_files and results:
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("✅ Berhasil Input", results['success'])
                        with col2:
                            st.metric("⏭️ Dilewati", results['skipped'])
                        with col3:
                            st.metric("❌ Gagal", results['failed'])
                    
                    st.balloons()  # Animasi celebrasi
                    st.rerun()  # Refresh untuk menampilkan area download
                    
                except Exception as e:
                    st.error(f"❌ **Error dalam proses otomatis:** {str(e)}")
                    
        else:
            st.info("📤 **Silakan upload file PDF Bank Statement untuk memulai proses otomatis**")
            
            # Info tentang proses
            with st.expander("ℹ️ Informasi Proses Otomatis"):
                st.markdown("""
                **Setelah upload file PDF (dan opsional Master Excel), sistem akan otomatis menjalankan:**
                
                1. 📄 **Ekstraksi PDF** - Mengambil data transaksi dari PDF Bank Statement
                2. 🔍 **Ekstraksi SETORTUNAI** - Memisahkan data rusun dan non-rusun
                3. 🛠️ **Filter Data** - Memfilter data lengkap dan tahun yang didukung
                4. 📊 **Ekstrak Master** - Mengambil data dari Master Excel (jika tersedia)
                5. 💰 **Kalkulasi Denda** - Menghitung denda berdasarkan selisih pembayaran
                6. 📝 **Input ke Master** - Memasukkan data ke Master Excel (jika tersedia)
                7. 📤 **Export Hasil** - Membuat file laporan status
                
                **Hasil yang didapat:**
                - File laporan status dengan 4 sheet (Status Input, Cek Manual, Kasda, Denda)
                - File Master Excel yang sudah terupdate (jika Master Excel diupload)
                
                **Waktu estimasi:** 30-60 detik tergantung ukuran file
                """)
            
            # Info file yang diperlukan
            with st.expander("📋 File yang Diperlukan"):
                st.markdown("""
                **Wajib:**
                - 📄 **PDF Bank Statement** - File PDF hasil download dari internet banking
                
                **Opsional (untuk fitur lengkap):**
                - 📊 **Master Excel 2024** - File master data untuk tahun 2024
                - 📊 **Master Excel 2025** - File master data untuk tahun 2025
                
                **Catatan:**
                - Jika Master Excel tidak diupload, sistem tetap bisa memproses data tetapi tidak akan ada input ke Master Excel
                - Hasil export tetap akan tersedia dengan data yang berhasil diekstrak
                """)
    
    # Footer info
    st.markdown("---")
    st.markdown("**💡 Tips:** Upload semua file sekaligus lalu klik tombol proses untuk hasil optimal!")

if __name__ == "__main__":
    main()
